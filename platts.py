#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : lusheng

from pylab import *
import matplotlib.pyplot as plt
from xlrd import open_workbook
from xlrd import xldate_as_tuple
from xlrd import xldate_as_datetime
import matplotlib.dates as mdates
import datetime
from bs4 import BeautifulSoup
import urllib.request
import re
import globalmap as gl
from openpyxl import load_workbook
from selenium import webdriver
from openpyxl import load_workbook
import openpyxl
from openpyxl.chart.axis import DateAxis
from openpyxl.chart import (
    LineChart,
    Reference,
)

# gl._init()

def platts(username,password):
    urlbase = 'http://www.f139.cn/ore/index.html'
    headers = {
        'Accept': 'text / html, application / xhtml + xml, application / xml;q = 0.9, image / webp, image / apng, * / *;q = 0.8',
        'Accept-Encoding':'gzip,deflate',
        'Accept-Language': 'zh-CN,zh;q = 0.9',
        'Cache-Control': 'max - age = 0',
        'Connection': 'keep - alive',
        'Cookie': 'cck_lasttime = 1536541528002;cck_count = 0;JSESSIONID = A839BFF65A229EC89A3269F123C9AE2E;Hm_lvt_e11e5fa7b1c17369dacfb3f063d64cae = 1534742448, 1535338601, 1535953657, 1536541537;Hm_lpvt_e11e5fa7b1c17369dacfb3f063d64cae = 1536541604',
        'User-Agent': 'Mozilla / 5.0(WindowsNT10.0;Win64;x64) AppleWebKit / 537.36(KHTML, likeGecko) Chrome / 68.0 3440.75 Safari / 537.36'
    }
    req = urllib.request.Request(urlbase, headers=headers)
    response1 = urllib.request.urlopen(req)
    html=response1.read().decode('utf-8')

    soup = BeautifulSoup(html, "html.parser")
    lists = soup.find_all('div', class_="w252 mod_convenient right")
    lists = str(lists)
    # print(lists)
    links = []
    links = re.findall(r"http://www.f139.cn/news/detail/([a-zA-Z0-9]+).html",lists)
    titles = re.findall(r"title=\"(.+?)\"",lists)
    urllist = {}
    urllist2 = {}
    i = len(links)-1
    while i >= 0:
        title = titles[i]
        link = 'http://www.f139.cn/news/detail/' + links[i] + '.html'
        # print(title, link)
        urllist[title]=link
        i -= 1
    for key in urllist:
        if '全'in key:
            urllist2[key] = str(urllist[key])
    print(urllist2)

    #新建一个文件用于保存数据
    # f = open('platts.txt', 'w', encoding='utf-8')
    excel_path = 'C:\\Users\\LUS\\Desktop\\周报材料\\周分析会议数据.xlsx'

    for k in range(len(urllist2)):
        # print(k)
        indexurl = list(urllist2.values())[k]
        indextitle = list(urllist2.keys())[k]
        print(indextitle[:-10], indexurl)
        options = webdriver.ChromeOptions()
        prefs = {"profile.managed_default_content_settings.images": 2}
        options.add_experimental_option("prefs", prefs)
        options.add_argument('--headless')
        browser = webdriver.Chrome(chrome_options=options)
        browser.get(indexurl)
        input_first = browser.find_element_by_tag_name('table')
        # print(input_first.text)
        # f.write(indextitle)
        # f.write(input_first.text)
        text_list =input_first.text.split()
        print(input_first.text.split())
        #获取需要写入excel表的数据
        write_data = [indextitle[:-10],text_list[10],text_list[4],text_list[7]]
        data1 = re.findall(r"([0-9]+)月", indextitle[:-10])
        data2 = re.findall(r"月([0-9]+)日", indextitle[:-10])
        # print(data1,data2)
        if len(data1[0]) == 1:
            data1[0] = '0' + data1[0]
        date = '2019-' + data1[0] + '-' + data2[0]
        print(date,text_list[10],text_list[4],text_list[7])
        # price_date= datetime.strptime(date, '%Y-%m-%d')
        # print(price_date)
        # f.write('\n\n')
        browser.quit()

        # 打开已经存在的表格并实例化，准备进行修改操作
        wb = load_workbook(excel_path)
        # print(wb.sheetnames)
        sheet = wb.get_sheet_by_name("普氏、MYSTEEL指数")
        # 查找需要修改的内容
        n_of_rows = sheet.max_row +1
        n_of_cols = sheet.max_column
        haddate=[]
        for i in range(1,n_of_rows):
            haddate.append(str(sheet.cell(row=i, column=1).value)[:10])
        print(n_of_rows,n_of_cols,haddate[-5:])
        # 写入数据参数包括行号、列号、和值（其中参数不止这些）
        # sheet["A%d" % n_of_rows].value = indextitle[:-10]

        if date in haddate:
            pass
        else:
            sheet["A%d" % n_of_rows].value = date
            sheet["A%d" % n_of_rows].number_format = 'yyyy-mm-dd'
            sheet["B%d" % n_of_rows].value = float(text_list[10])
            sheet["C%d" % n_of_rows].value = float(text_list[4])
            sheet["G%d" % n_of_rows].value = float(text_list[7])
        wb.save('C:\\Users\\LUS\\Desktop\\周报材料\\周分析会议数据.xlsx')

    excel_path = 'C:\\Users\\LUS\\Desktop\\周报材料\\周分析会议数据.xlsx'
    wb = load_workbook(excel_path)
    ws = wb.get_sheet_by_name("普氏、MYSTEEL指数")
    n_of_rows = ws.max_row - 250
    # global date1, date1_62, date1_58, date2, date2_62, date2_58
    # date1 = ws['A%d' % ws.max_row].value
    # date1_62 = ws['B%d'% ws.max_row].value
    # date1_58 = ws['C%d' % ws.max_row].value
    # date2 = ws['A%d' % (ws.max_row - 5)].value
    # date2_62 = ws['B%d' % (ws.max_row - 5)].value
    # date2_58 = ws['C%d' % (ws.max_row - 5)].value

    gl.set_value('date1', ws['A%d' % ws.max_row].value)
    gl.set_value('date1_62', ws['B%d'% ws.max_row].value)
    gl.set_value('date1_58', ws['C%d' % ws.max_row].value)
    gl.set_value('date2', ws['A%d' % (ws.max_row - 5)].value)
    gl.set_value('date2_62', ws['B%d' % (ws.max_row - 5)].value)
    gl.set_value('date2_58', ws['C%d' % (ws.max_row - 5)].value)

    # data1 = Reference(ws, min_col=2, min_row=n_of_rows, max_row=ws.max_row)#数据来源，多系列可以一个个添加，否则标题不好处理
    # data2 = Reference(ws, min_col=3, min_row=n_of_rows, max_row=ws.max_row)
    # se1 = openpyxl.chart.Series(data1, title=ws['B1'].value)
    # se2 = openpyxl.chart.Series(data2, title=ws['C1'].value)
    # se2.graphicalProperties.line.solidFill = "00AAAA"#设置系列数据线条颜色
    # c2 = LineChart()
    # c2.title = "普氏58、62指数"
    # c2.style = 12
    # c2.y_axis.title = "美元"
    # c2.y_axis.crossAx = 500
    # c2.x_axis = DateAxis(crossAx=100)
    # c2.x_axis.number_format = 'mmm-e' #日期坐标轴的显示格式
    # c2.x_axis.majorTimeUnit = "months" #日期坐标轴间隔单位
    # c2.y_axis.scaling.min = 10  #y坐标的区间
    # c2.y_axis.scaling.max = 100
    # c2.append(se1)#向图表中增加数据系列
    # c2.append(se2)
    # dates = Reference(ws, min_col=1, min_row=n_of_rows, max_row=ws.max_row)   #日期坐标轴数据
    # c2.set_categories(dates)
    # c2.height = 11#图表大小
    # c2.width = 18
    # ws.add_chart(c2, "F%d" % (ws.max_row - 20))   #添加图表到excel表指定位置
    # wb.save("C:\\Users\\LUS\\Desktop\\周分析会议数据2.xlsx")
    # print('platts指数图表绘制完成')
    wb.close()

    x_data = []
    y_data1 = []
    y_data2 = []
    mpl.rcParams['font.sans-serif'] = ['SimHei']
    mpl.rcParams['axes.unicode_minus'] = False
    wb = open_workbook('C:\\Users\\LUS\\Desktop\\周报材料\\周分析会议数据.xlsx')
    for s in wb.sheets():
        # print('Sheet:',s.name )
        for row in range(s.nrows - 250, s.nrows):
            # print('the row is:',row )
            if type(s.cell_value(row, 0)) is str:
                date = s.cell_value(row, 0)
            else:
                date = xldate_as_datetime(s.cell_value(row, 0), 0)
                date = str(date.strftime('%Y-%m-%d'))
            print(date)
            # values.append(date)
            x_data.append(date)
            y_data1.append(s.cell_value(row, 1))
            y_data2.append(s.cell_value(row, 2))
    x_data = [datetime.datetime.strptime(str(d), '%Y-%m-%d').date() for d in x_data]
    print(x_data)
    print(y_data1)
    print(y_data2)
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    plt.gca().xaxis.set_major_locator(mdates.DayLocator())

    fig = plt.figure()
    ax1 = fig.add_subplot(111)
    plt.plot(x_data, y_data1, label=u"62指数", linewidth=3)
    plt.plot(x_data, y_data2, label=u"58指数", linewidth=3)
    plt.title("普氏58、62指数")
    plt.legend()

    plt.gcf().autofmt_xdate()  # 自动旋转日期标记
    month = mdates.MonthLocator()  # 主刻度为每月
    ax1.xaxis.set_major_locator(month)  # 设置主刻度
    plt.ylim(10, 100)
    plt.grid(axis="y")
    # plt.xlabel("date")
    plt.ylabel("美元")
    plt.show()
    fig.savefig('C:\\Users\\LUS\Desktop\\周报材料\\普氏指数.png')
    print('普氏指数作图完成')


# platts('','')
